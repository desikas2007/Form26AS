import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def style_worksheet(ws, has_index=False):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = ws[2]


def save_to_excel(tables, fields, page_timing, summary, output_dir="output"):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"26AS_Extraction_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        
        pd.DataFrame([summary]).to_excel(writer, sheet_name="SUMMARY", index=False)
        
        summary_ws = writer.sheets["SUMMARY"]
        style_worksheet(summary_ws)
        
        pd.DataFrame(page_timing).to_excel(writer, sheet_name="PAGE_TIMING", index=False)
        timing_ws = writer.sheets["PAGE_TIMING"]
        style_worksheet(timing_ws)
        
        pd.DataFrame(list(fields.items()), columns=["Field", "Value"])\
            .to_excel(writer, sheet_name="FORM_FIELDS", index=False)
        fields_ws = writer.sheets["FORM_FIELDS"]
        style_worksheet(fields_ws)
        
        for i, df in enumerate(tables, 1):
            sheet_name = f"Table_{i}"
            if len(sheet_name) > 31:
                sheet_name = f"T_{i}"
            
            df.to_excel(writer, sheet_name=sheet_name, index=True)
            ws = writer.sheets[sheet_name]
            style_worksheet(ws, has_index=True)
            
            for col_idx, col_name in enumerate(df.columns, start=2):
                cell = ws.cell(row=1, column=col_idx)
                if col_idx == 2:
                    cell.value = "S.No"
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    return output_path


def create_combined_sheet(output_path, combined_data, sheet_name="COMBINED_DATA"):
    wb = load_workbook(output_path)
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    start_row = 1
    for df in combined_data:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        start_row += len(df) + 3
    
    style_worksheet(ws, has_index=True)
    wb.save(output_path)


def export_to_csv(tables, fields, page_timing, summary, output_dir="output_csv"):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"26AS_Extraction_{timestamp}"
    
    pd.DataFrame([summary]).to_csv(f"{output_dir}/{base_name}_summary.csv", index=False)
    
    pd.DataFrame(page_timing).to_csv(f"{output_dir}/{base_name}_page_timing.csv", index=False)
    
    pd.DataFrame(list(fields.items()), columns=["Field", "Value"])\
        .to_csv(f"{output_dir}/{base_name}_fields.csv", index=False)
    
    for i, df in enumerate(tables, 1):
        df.to_csv(f"{output_dir}/{base_name}_table_{i}.csv", index=True)
    
    return output_dir
